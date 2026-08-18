"""
EXCHANGE_RATE.PY
================
Gets the current exchange rate from the GetEXRate SOAP service.

CACHING: the rate only actually changes twice a day (see
config.EXRATE_REFRESH_TIMES, default 10:00 and 14:00). The SOAP service is
only called when the clock has crossed one of those times since the last
successful call ("we've entered a new rate slot") — every other request
(a file can be converted any number of times between refreshes) reuses the
cached value instead of hitting the service again.

Every fetched rate is appended as a row to config.EXRATE_HISTORY_FILE
(an .xlsx). That file also doubles as the persisted cache: on startup, if
its last row already belongs to the current slot, that rate is reused
immediately instead of forcing an extra SOAP call right at process start.

If the service call fails for any reason (no internet, server error,
bad response), this falls back to:
    1. The last rate that worked successfully (this run, or restored from
       the history file), or
    2. config.EXRATE_DEFAULT_FALLBACK_RATE (= "1") if nothing has worked yet.
"""

import json
import logging
import os
import re
import time
import uuid
from datetime import datetime, timedelta, time as dtime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

import requests
from openpyxl import Workbook, load_workbook

import config

logger = logging.getLogger(__name__)

# In-memory cache. _cached_slot_key identifies which refresh slot
# _cached_rate belongs to (e.g. "2026-08-01 10:00") — see _current_slot_key.
_cached_rate = None
_cached_slot_key = None

# Remembers the last rate that worked at all (any slot), as a fallback of
# last resort if a fresh SOAP call fails and we need SOMETHING to return.
_last_successful_rate = None

# Whether we've already tried to restore the cache from the history file
# this run (only needs to happen once, on the first call).
_restored_from_history = False

# Human-readable detail on why the most recent SOAP call/parse failed, so
# the JSON call log can record something more useful than just "failed".
_last_call_error_detail = None

_HISTORY_HEADERS = ["Timestamp", "Slot", "Rate", "Source", "Note"]

_SOAP_REQUEST_TEMPLATE = """<?xml version="1.0" encoding="utf-8"?>
<soap:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/">
  <soap:Body>
    <GetEXRate xmlns="ClientWebService">
      <AGENT_CODE>{agent_code}</AGENT_CODE>
      <USER_ID>{user_id}</USER_ID>
      <AGENT_SESSION_ID>{agent_session_id}</AGENT_SESSION_ID>
      <TRANSFER_AMOUNT>{transfer_amount}</TRANSFER_AMOUNT>
      <PAYMENT_MODE>{payment_mode}</PAYMENT_MODE>
      <CALC_BY>{calc_by}</CALC_BY>
      <LOCATION_ID>{location_id}</LOCATION_ID>
      <PAYOUT_COUNTRY>{payout_country}</PAYOUT_COUNTRY>
      <SIGNATURE>{signature}</SIGNATURE>
    </GetEXRate>
  </soap:Body>
</soap:Envelope>"""

_EXCHANGE_RATE_PATTERN = re.compile(r"<EXCHANGE_RATE>(.*?)</EXCHANGE_RATE>", re.IGNORECASE | re.DOTALL)
_CODE_PATTERN = re.compile(r"<CODE>(.*?)</CODE>", re.IGNORECASE | re.DOTALL)
_MESSAGE_PATTERN = re.compile(r"<MESSAGE>(.*?)</MESSAGE>", re.IGNORECASE | re.DOTALL)


# ---------------------------------------------------------------------------
# REFRESH SLOTS — "which of today's (or the latest) scheduled rate refreshes
# are we currently under?"
# ---------------------------------------------------------------------------

def _parse_refresh_times() -> list:
    """Parse config.EXRATE_REFRESH_TIMES (["10:00", "14:00"]) into sorted
    datetime.time objects."""
    parsed = []
    for value in config.EXRATE_REFRESH_TIMES:
        hour_str, minute_str = value.split(":")
        parsed.append(dtime(int(hour_str), int(minute_str)))
    return sorted(parsed)


def _current_slot_key(now: datetime) -> str:
    """Identify the current rate-refresh slot as a stable string, e.g.
    "2026-08-01 10:00". Times before the first refresh of the day belong
    to the previous day's last slot (yesterday's last-fetched rate still
    holds until today's first scheduled refresh)."""
    refresh_times = _parse_refresh_times()
    today = now.date()

    todays_slot_datetimes = [datetime.combine(today, t) for t in refresh_times]
    passed_today = [dt for dt in todays_slot_datetimes if dt <= now]

    if passed_today:
        slot_dt = max(passed_today)
    else:
        yesterday = today - timedelta(days=1)
        slot_dt = datetime.combine(yesterday, refresh_times[-1])

    return slot_dt.strftime("%Y-%m-%d %H:%M")


# ---------------------------------------------------------------------------
# HISTORY FILE (persisted cache + audit trail)
# ---------------------------------------------------------------------------

def _read_last_successful_history_row():
    """Scan the history file backwards for the most recent row whose
    source was an actual successful SOAP call ("soap") — deliberately
    ignores fallback rows (fallback_last_successful / fallback_default),
    since those don't represent a validated rate for that slot and
    shouldn't be treated as "we already have this slot covered" on
    restart. Returns (slot_key, rate) or (None, None)."""
    if not config.EXRATE_HISTORY_FILE.exists():
        return None, None
    try:
        wb = load_workbook(config.EXRATE_HISTORY_FILE, read_only=True)
        ws = wb.active
        rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0] is not None]
        wb.close()

        for row in reversed(rows):
            # columns: Timestamp, Slot, Rate, Source, Note
            slot_key, rate_str, source = row[1], row[2], row[3]
            if source != "soap":
                continue
            try:
                return slot_key, Decimal(str(rate_str))
            except (InvalidOperation, TypeError):
                continue
        return None, None
    except Exception:
        logger.exception("Could not read exchange rate history file; ignoring it.")
        return None, None


def _append_history_row(slot_key: str, rate: Decimal, source: str, note: str = "") -> None:
    """Append one row to the history .xlsx, creating it with a header row
    if it doesn't exist yet."""
    try:
        if config.EXRATE_HISTORY_FILE.exists():
            wb = load_workbook(config.EXRATE_HISTORY_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Rate History"
            ws.append(_HISTORY_HEADERS)

        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            slot_key,
            str(rate),
            source,
            note,
        ])
        wb.save(config.EXRATE_HISTORY_FILE)
    except Exception:
        # Never let a logging failure break rate fetching itself.
        logger.exception("Could not write to exchange rate history file.")


def _restore_cache_from_history_if_needed(required_slot: str) -> None:
    """On the first call in this run, try to seed the in-memory cache from
    the history file so a process restart doesn't force an unnecessary
    SOAP call if today's slot rate was already fetched by an earlier run.

    Only ever restores from a row whose source was an actual successful
    SOAP call — a fallback value (used because the service was
    unreachable, e.g. an empty/misconfigured config.py) must NOT get
    "locked in" as if it were a validated rate for that slot, or a
    restart would keep reusing the fallback forever instead of retrying
    the real service once it's fixed."""
    global _cached_rate, _cached_slot_key, _last_successful_rate, _restored_from_history

    if _restored_from_history:
        return
    _restored_from_history = True

    slot_key, rate = _read_last_successful_history_row()
    if rate is None:
        return

    _last_successful_rate = rate
    if slot_key == required_slot:
        _cached_rate = rate
        _cached_slot_key = slot_key
        logger.info(
            "Restored exchange rate %s for slot %s from history file "
            "(no SOAP call needed at startup).", rate, slot_key
        )


# ---------------------------------------------------------------------------
# SOAP CALL
# ---------------------------------------------------------------------------

def _build_request_body() -> str:
    """Fill in the SOAP request template with values from config.py."""
    return _SOAP_REQUEST_TEMPLATE.format(
        agent_code=config.EXRATE_AGENT_CODE,
        user_id=config.EXRATE_USER_ID,
        agent_session_id=config.EXRATE_AGENT_SESSION_ID,
        transfer_amount=config.EXRATE_TRANSFER_AMOUNT,
        payment_mode=config.EXRATE_PAYMENT_MODE,
        calc_by=config.EXRATE_CALC_BY,
        location_id=config.EXRATE_LOCATION_ID,
        payout_country=config.STATIC_VALUES["PayoutCountry"],
        signature=config.EXRATE_SIGNATURE,
    )


def _call_soap_service():
    """Call the GetEXRate SOAP endpoint. Returns the raw XML text, or None
    if the call failed for ANY reason — including a bad/incomplete
    config.py (missing values, etc), not just a network error. Building
    the request body is deliberately inside this try/except: an
    incomplete config used to raise uncaught here and crash before any
    fallback or logging ever ran."""
    global _last_call_error_detail

    try:
        body = _build_request_body()
        headers = {
            "Content-Type": "text/xml; charset=utf-8",
            "SOAPAction": f'"{config.EXRATE_SOAP_ACTION}"',
        }
        response = requests.post(
            config.EXRATE_SOAP_URL,
            data=body.encode("utf-8"),
            headers=headers,
            timeout=config.EXRATE_REQUEST_TIMEOUT_SECONDS,
        )
    except requests.RequestException as e:
        logger.exception("Exchange rate request failed (network error)")
        _last_call_error_detail = f"network error: {e}"
        return None
    except Exception as e:
        logger.exception(
            "Exchange rate request failed (likely an incomplete config.py "
            "— check EXRATE_* settings and STATIC_VALUES)."
        )
        _last_call_error_detail = f"config/request build error: {e}"
        return None

    if response.status_code != 200:
        logger.error("Exchange rate service returned status %d", response.status_code)
        _last_call_error_detail = f"HTTP status {response.status_code}"
        return None

    return response.text


def _parse_rate(xml_text: str):
    """Pull the EXCHANGE_RATE value out of the SOAP response. Returns a
    Decimal, or None if the response was an error or unreadable."""
    global _last_call_error_detail

    code_match = _CODE_PATTERN.search(xml_text)
    if code_match:
        code = code_match.group(1).strip()
        if code not in ("0", "00", "000", "1", ""):
            message_match = _MESSAGE_PATTERN.search(xml_text)
            message = message_match.group(1).strip() if message_match else ""
            logger.error("Exchange rate service error CODE=%s MESSAGE=%s", code, message)
            _last_call_error_detail = f"service error CODE={code} MESSAGE={message}"
            return None

    rate_match = _EXCHANGE_RATE_PATTERN.search(xml_text)
    if not rate_match:
        logger.error("Exchange rate response missing EXCHANGE_RATE field")
        _last_call_error_detail = "response missing EXCHANGE_RATE field"
        return None

    try:
        raw_rate = Decimal(rate_match.group(1).strip())
    except InvalidOperation:
        logger.error("Exchange rate response had a non-numeric rate")
        _last_call_error_detail = f"non-numeric rate: {rate_match.group(1).strip()!r}"
        return None

    # Round to exactly 2 decimal places HERE, once, at the source — not
    # at the point of use in converter.py. The raw SOAP response can
    # carry more precision than the rate you'd see quoted elsewhere
    # (e.g. "151.9100" on a portal vs. a longer raw value underneath),
    # and without normalizing it here, two separate fetches that landed
    # a hair apart could each cache a slightly different high-precision
    # value — meaning different rows/files could silently divide by
    # slightly different rates even though they're "the same" 151.91.
    # Rounding once at fetch time guarantees every row in every file
    # divides by the exact same 2-decimal rate for the whole slot.
    rate = raw_rate.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if rate != raw_rate:
        logger.info("Exchange rate rounded from %s to %s (2 decimal places).", raw_rate, rate)
    return rate


# ---------------------------------------------------------------------------
# JSON CALL LOG (independent of the .xlsx history — always written)
# ---------------------------------------------------------------------------

def _append_json_call_log(status: str, rate, error_detail: str = "") -> None:
    """Append one entry to exchange_rate_log.json recording whether this
    SOAP call attempt succeeded or failed. Written with an atomic
    temp-file + os.replace, same pattern as alerts.py, so a crash or a
    concurrent read never sees a half-written file. This is deliberately
    separate from the .xlsx history so you always get a record of call
    attempts even if the xlsx write has a problem."""
    path = config.EXRATE_JSON_LOG_FILE
    entry = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "status": status,  # "success" or "failure"
        "rate": str(rate) if rate is not None else None,
        "error": error_detail or None,
    }

    entries = []
    if path.exists():
        for _ in range(3):
            try:
                with path.open("r", encoding="utf-8") as f:
                    entries = json.load(f)
                break
            except Exception:
                time.sleep(0.1)
        if not isinstance(entries, list):
            entries = []

    entries.append(entry)
    if len(entries) > config.EXRATE_JSON_LOG_MAX_ENTRIES:
        entries = entries[-config.EXRATE_JSON_LOG_MAX_ENTRIES:]

    tmp_path = path.with_name(f"{path.name}.{uuid.uuid4().hex}.tmp")
    for _ in range(3):
        try:
            with tmp_path.open("w", encoding="utf-8") as f:
                json.dump(entries, f, indent=2)
            os.replace(tmp_path, path)  # atomic on both Windows and POSIX
            return
        except Exception:
            time.sleep(0.1)

    logger.exception("Could not write exchange rate JSON call log.")
    try:
        if tmp_path.exists():
            tmp_path.unlink()
    except Exception:
        pass


# ---------------------------------------------------------------------------
# PUBLIC ENTRY POINT
# ---------------------------------------------------------------------------

def get_exchange_rate() -> Decimal:
    """Get the current exchange rate. Public entry point — wraps the real
    logic in a catch-all so that ANY unexpected failure (bad config,
    bug, etc.) still logs to the JSON call log and still returns a usable
    rate, instead of crashing the file conversion that called this."""
    try:
        return _get_exchange_rate_inner()
    except Exception as e:
        logger.exception("Unexpected error getting exchange rate.")
        fallback = _last_successful_rate or Decimal(config.EXRATE_DEFAULT_FALLBACK_RATE)
        _append_json_call_log("failure", fallback, error_detail=f"unexpected error: {e}")
        return fallback


def _get_exchange_rate_inner() -> Decimal:
    """Get the current exchange rate.

    Only calls the live SOAP service if the clock has moved into a new
    refresh slot (config.EXRATE_REFRESH_TIMES) since the last successful
    fetch; otherwise returns the cached rate straight away. On a SOAP
    failure, falls back to the last rate that worked (this run, or
    restored from the history file), or to
    config.EXRATE_DEFAULT_FALLBACK_RATE if nothing has worked yet.
    """
    global _cached_rate, _cached_slot_key, _last_successful_rate, _last_call_error_detail

    now = datetime.now()
    required_slot = _current_slot_key(now)

    _restore_cache_from_history_if_needed(required_slot)

    if _cached_rate is not None and _cached_slot_key == required_slot:
        logger.info(
            "Using cached exchange rate %s (slot %s) — no SOAP call needed.",
            _cached_rate, required_slot
        )
        return _cached_rate

    logger.info("Refreshing exchange rate for new slot %s...", required_slot)
    _last_call_error_detail = None
    xml_response = _call_soap_service()

    rate = None
    if xml_response is not None:
        rate = _parse_rate(xml_response)

    if rate is not None:
        _cached_rate = rate
        _cached_slot_key = required_slot
        _last_successful_rate = rate
        _append_history_row(required_slot, rate, source="soap")
        _append_json_call_log("success", rate)
        return rate

    # SOAP call failed — DON'T update _cached_slot_key, so the next call
    # (even within the same slot) retries the SOAP service instead of
    # silently sticking with a failure.
    error_detail = _last_call_error_detail or "unknown error"

    if _last_successful_rate is not None:
        logger.warning(
            "Exchange rate refresh failed for slot %s; using last "
            "successful rate: %s", required_slot, _last_successful_rate
        )
        _append_history_row(
            required_slot, _last_successful_rate, source="fallback_last_successful",
            note="SOAP call failed for this slot"
        )
        _append_json_call_log(
            "failure", _last_successful_rate,
            error_detail=f"{error_detail} (used last successful rate as fallback)"
        )
        return _last_successful_rate

    fallback = Decimal(config.EXRATE_DEFAULT_FALLBACK_RATE)
    logger.warning("No exchange rate available yet; using default: %s", fallback)
    _append_history_row(
        required_slot, fallback, source="fallback_default",
        note="SOAP call failed and no prior successful rate this run"
    )
    _append_json_call_log(
        "failure", fallback,
        error_detail=f"{error_detail} (no prior rate — used default fallback)"
    )
    return fallback
